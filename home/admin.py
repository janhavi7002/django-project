from django.contrib import admin
from django.utils import timezone
from django.db import models
from django.forms import TextInput
from django.http import HttpResponse
from django.urls import path

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

from .models import Student, MaintenanceHistory


@admin.register(Student)
class StudentAdmin(admin.ModelAdmin):

    # ---------------------------------------------------------
    # CUSTOM EXCEL BUTTON
    # ---------------------------------------------------------

    change_list_template = "admin/home/student/change_list.html"

    def get_urls(self):
        urls = super().get_urls()

        custom_urls = [
            path(
                "export-excel/",
                self.admin_site.admin_view(self.export_excel),
                name="home_student_export_excel",
            ),
        ]

        return custom_urls + urls

    def export_excel(self, request):
        """
        Export the currently displayed/filtered students to Excel.
        """

        # Get the current ChangeList.
        # This means search and filters are respected.
        changelist = self.get_changelist_instance(request)
        queryset = changelist.get_queryset(request)

        # Create Excel workbook
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = "Students"

        # -----------------------------------------------------
        # EXCEL HEADERS
        # -----------------------------------------------------

        headers = [
            "Name",
            "Email",
            "Admission Number",
            "Branch",
            "Year",
            "Contact Number",
            "Component Name",
            "Quantity",
            "Status",
            "Component Issue Date",
            "Component Due Date",
            "Faculty Referred",
            "Deleted",
            "Remarks",
        ]

        worksheet.append(headers)

        # -----------------------------------------------------
        # STYLE HEADER
        # -----------------------------------------------------

        header_fill = PatternFill(
            fill_type="solid",
            fgColor="217346"
        )

        header_font = Font(
            bold=True,
            color="FFFFFF"
        )

        header_alignment = Alignment(
            horizontal="center",
            vertical="center"
        )

        for cell in worksheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # -----------------------------------------------------
        # ADD STUDENT DATA
        # -----------------------------------------------------

        for student in queryset:

            row = [
                student.name,
                student.email,
                student.admission_number,
                student.branch,
                student.year,
                student.contact_number,
                student.component_name,
                student.quantity,
                student.status,
                student.componentissue_date,
                student.componentdue_date,
                student.faculty_referred,
                "Yes" if student.is_deleted else "No",
                student.remarks,
            ]

            worksheet.append(row)

        # -----------------------------------------------------
        # FORMAT CELLS
        # -----------------------------------------------------

        for row in worksheet.iter_rows():

            for cell in row:
                cell.alignment = Alignment(
                    vertical="center"
                )

        # -----------------------------------------------------
        # AUTO COLUMN WIDTH
        # -----------------------------------------------------

        for column_cells in worksheet.columns:

            max_length = 0

            column_letter = get_column_letter(
                column_cells[0].column
            )

            for cell in column_cells:

                try:
                    cell_length = len(str(cell.value))

                    if cell_length > max_length:
                        max_length = cell_length

                except Exception:
                    pass

            # Limit maximum width
            adjusted_width = min(
                max_length + 2,
                40
            )

            worksheet.column_dimensions[
                column_letter
            ].width = adjusted_width

        # -----------------------------------------------------
        # FREEZE HEADER ROW
        # -----------------------------------------------------

        worksheet.freeze_panes = "A2"

        # -----------------------------------------------------
        # ENABLE FILTER IN EXCEL
        # -----------------------------------------------------

        if worksheet.max_row > 1:
            worksheet.auto_filter.ref = worksheet.dimensions

        # -----------------------------------------------------
        # CREATE RESPONSE
        # -----------------------------------------------------

        response = HttpResponse(
            content_type=(
                "application/vnd.openxmlformats-officedocument."
                "spreadsheetml.sheet"
            )
        )

        response["Content-Disposition"] = (
            'attachment; filename="students.xlsx"'
        )

        workbook.save(response)

        return response

    # ---------------------------------------------------------
    # STUDENT LIST DISPLAY
    # ---------------------------------------------------------

    list_display = (
        'name',
        'email',
        'admission_number',
        'branch',
        'year',
        'contact_number',
        'component_name',
        'quantity',
        'status',
        'componentissue_date',
        'componentdue_date',
        'faculty_referred',
        'is_deleted',
        'remarks',
    )

    # ---------------------------------------------------------
    # EDITABLE FIELD
    # ---------------------------------------------------------

    list_editable = (
        'remarks',
    )

    # ---------------------------------------------------------
    # TEXT FIELD STYLE
    # ---------------------------------------------------------

    formfield_overrides = {
        models.TextField: {
            'widget': TextInput(
                attrs={
                    'style': (
                        'width:150px; '
                        'height:20px;'
                    )
                }
            )
        }
    }

    # ---------------------------------------------------------
    # FILTERS
    # ---------------------------------------------------------

    list_filter = (
        'status',
        'branch',
        'year',
        'is_deleted',
    )

    # ---------------------------------------------------------
    # SEARCH
    # ---------------------------------------------------------

    search_fields = (
        'name',
        'email',
        'admission_number',
        'component_name',
    )

    # ---------------------------------------------------------
    # ADMIN ACTIONS
    # ---------------------------------------------------------

    actions = [
        'move_to_history',
        'restore_records',
    ]

    # ---------------------------------------------------------
    # MOVE TO MAINTENANCE HISTORY
    # ---------------------------------------------------------

    @admin.action(
        description="Move selected records to Maintenance History"
    )
    def move_to_history(self, request, queryset):

        for obj in queryset:

            MaintenanceHistory.objects.create(
                student_name=obj.name,
                admission_number=obj.admission_number,
                component_name=obj.component_name,
                quantity=obj.quantity,
                deleted_by=request.user.username,
                reason="Deleted by Admin"
            )

            obj.is_deleted = True
            obj.deleted_at = timezone.now()

            obj.save()

    # ---------------------------------------------------------
    # RESTORE RECORDS
    # ---------------------------------------------------------

    @admin.action(
        description="Restore selected records"
    )
    def restore_records(self, request, queryset):

        queryset.update(
            is_deleted=False,
            deleted_at=None
        )


# -------------------------------------------------------------
# MAINTENANCE HISTORY
# -------------------------------------------------------------

@admin.register(MaintenanceHistory)
class MaintenanceHistoryAdmin(admin.ModelAdmin):

    list_display = (
        'student_name',
        'admission_number',
        'component_name',
        'quantity',
        'deleted_by',
        'reason',
    )